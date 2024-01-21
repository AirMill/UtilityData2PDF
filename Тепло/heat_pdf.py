import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from openpyxl import load_workbook
from datetime import datetime
from reportlab.pdfgen import canvas as pdf_canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import os
import json
from math import ceil

# Load custom font for PDF
pdfmetrics.registerFont(TTFont('ArialUnicode', 'arial.ttf'))

# Loading data from Excel


def load_data(file_path, sheet_name='Sheet1'):
    try:
        workbook = load_workbook(file_path)
        sheet = workbook[sheet_name]
        data = [(sheet.cell(row=row, column=1).value, sheet.cell(row=row, column=2).value, sheet.cell(
            row=row, column=3).value, sheet.cell(row=row, column=4).value) for row in range(2, sheet.max_row + 1)]
        return data
    except FileNotFoundError:
        messagebox.showerror("File Not Found", f"The Excel file '{
                             file_path}' is not found.")
        exit()


def create_widgets(frame, row, value1, value2, value3, value4, saved_values1, saved_values2, saved_values3):
    label_text = f"{
        value1} - {value2} - {value4}" if value1 is not None else ''
    label = ttk.Label(frame, text=label_text, font='Arial 11 bold')
    label.grid(row=row, column=0, padx=5, pady=5, sticky='w')

    # Entry 1
    entry_var1 = tk.StringVar(value=saved_values1.get(
        f"{value1}_{value2}_{value4}", ""))
    entry_vars.append(entry_var1)
    is_preloaded1 = entry_var1.get() != ""
    apply_entry_styling(frame, row, entry_var1, is_preloaded1, 1)

    label_text_comments = f"{value3}" if value3 is not None else ''
    label_comments = ttk.Label(
        frame, text=label_text_comments, font='Arial 10 bold')
    label_comments.grid(row=row, column=2, padx=5, pady=5, sticky='w')

    label2_text = "Корректировка на 24 часа:"
    label2 = ttk.Label(frame, text=label2_text, font='Arial 9 bold')
    label2.grid(row=row, column=3, padx=5, pady=5, sticky='w')

    # Entry 2
    entry_var2 = tk.StringVar(value=saved_values2.get(
        f"{value1}_{value2}_{value4}", ""))
    entry2_vars.append(entry_var2)
    is_preloaded2 = entry_var2.get() != ""
    apply_entry_styling(frame, row, entry_var2, is_preloaded2, 4)

    label3_text = "Потери Газпромэнерго:"
    label3 = ttk.Label(frame, text=label3_text, font='Arial 9 bold')
    label3.grid(row=row, column=5, padx=5, pady=5, sticky='w')

    # Entry 3
    entry_var3 = tk.StringVar(value=saved_values3.get(
        f"{value1}_{value2}_{value4}", ""))
    entry3_vars.append(entry_var3)
    is_preloaded3 = entry_var3.get() != ""
    apply_entry_styling(frame, row, entry_var3, is_preloaded3, 6)


def apply_entry_styling(frame, row, entry_var, is_preloaded, column):
    entry_bg_color = 'lightgrey' if is_preloaded else 'white'
    entry_font = ('Arial', 11, 'bold') if not is_preloaded else (
        'Arial', 10, 'italic')

    entry = ttk.Entry(frame, textvariable=entry_var,
                      font=entry_font, background=entry_bg_color)
    entry.grid(row=row, column=column, padx=5, pady=5, sticky='w')
    entry_widgets.append(entry)

    # Bind an event to change the background color when clicked
    entry.bind('<Button-1>', lambda event, entry=entry,
               is_preloaded=is_preloaded: on_entry_click(event, entry, is_preloaded))


def on_entry_click(event, entry, is_preloaded):
    if is_preloaded:
        # If preloaded, set the font to Arial 11 bold
        entry.config(font='Arial 11 bold', background='white')
        entry.icursor(0)  # Clear the italic styling
    else:
        # If not preloaded, toggle between bold and normal font
        current_font = entry.cget("font")
        new_font = "Arial 11 bold" if "bold" not in current_font else "Arial 11"
        entry.config(font=new_font)


# To use this function, you can call it in your loop like this:
# for i, (value1, value2, value3, value4) in enumerate(data):
#     create_widgets_extended(frame_middle, i, value1, value2, value3, value4, saved_values1, saved_values2, saved_values3)


def update_gui():
    global entry_vars, entry2_vars, entry3_vars, entry_widgets

    # Clear references to old widgets
    entry_vars = []
    entry2_vars = []
    entry3_vars = []
    entry_widgets = []

    # Load data from Excel file
    data = load_data('res/счетчики_тепло.xlsx')

    # Destroy existing widgets in the frame
    for widget in frame_middle.winfo_children():
        widget.destroy()

    # Create new widgets based on the updated data
    for i, (value1, value2, value3, value4) in enumerate(data):
        create_widgets(frame_middle, i, value1, value2, value3,
                       value4, saved_values1, saved_values2, saved_values3)


# submit functions

def on_submit():
    # Generate PDF
    generate_pdf('test.pdf')

    # Save values to JSON files
    save_values_to_file('saved_values1.json', {f"{value1}_{value2}_{value4}": entry_var.get()
                                               for entry_var, (value1, value2, _, value4) in zip(entry_vars, data)})

    save_values_to_file('saved_values2.json', {f"{value1}_{value2}_{value4}": entry2_var.get()
                                               for entry2_var, (value1, value2, _, value4) in zip(entry2_vars, data)})

    save_values_to_file('saved_values3.json', {f"{value1}_{value2}_{value4}": entry3_var.get()
                                               for entry3_var, (value1, value2, _, value4) in zip(entry3_vars, data)})

    # Disable entry widgets after generating PDF
    for entry in entry_widgets:
        entry.config(state="disabled")

    # Change button style to red
    button_style.configure(
        'Submit.TButton', background='red', foreground='black')

    # Change button text and command based on the current state
    if button_state.get() == 'submit':
        submit_button.config(text="Выход", command=window.destroy)
        button_state.set('exit')
    else:
        submit_button.config(text="Отправить", command=on_submit)


def generate_pdf(filename):
    current_datetime = datetime.now().strftime("%d%m%Y")
    pdf_filename = f"Показания_теплоснабжение_{current_datetime}.pdf"

    pdf_canvas_obj = pdf_canvas.Canvas(pdf_filename)

    # Use Arial Unicode MS font
    pdf_canvas_obj.setFont("ArialUnicode", 16)

    # Header
    pdf_canvas_obj.drawCentredString(
        300, 750, f'Показания теплоснабжения за: {datetime.now().strftime("%B %Y")}')

    pdf_canvas_obj.setFont("ArialUnicode", 15)
    pdf_canvas_obj.drawCentredString(
        300, 810, f"Протокол показаний счетчиков теплоснабжения. ")
    pdf_canvas_obj.drawCentredString(
        300, 780, f"Адрес: Мос. обл., Солнечногорский р-н., пос. Поварово, мкр. Поваровка-12")

    entries_per_page = 10  # Adjust as needed
    total_entries = len(entry_vars)

    for page_number in range(ceil(total_entries / entries_per_page)):
        start_index = page_number * entries_per_page
        end_index = min((page_number + 1) * entries_per_page, total_entries)

        # Content
        y = 700  # Starting y-coordinate
        for i in range(start_index, end_index):
            entry_var, entry2_var, entry3_var = entry_vars[i], entry2_vars[i], entry3_vars[i]
            value1, value2, value3, value4 = data[i]

            user_input = entry_var.get() or "Нет показаний"
            user_input2 = entry2_var.get() or ""
            user_input3 = entry3_var.get() or ""

            pdf_canvas_obj.setFont("ArialUnicode", 12)
            pdf_canvas_obj.drawString(
                100, y, f"Показания счетчика - {value1} {value2} {value4} - : {user_input}")
            y -= 20  # Adjust the vertical spacing
            pdf_canvas_obj.drawString(
                120, y, f"Корректировка на 24 часа: {user_input2}")
            y -= 20  # Adjust the vertical spacing
            pdf_canvas_obj.drawString(
                120, y, f"Потери Газпромэнерго: {user_input3}")
            y -= 20  # Adjust the vertical spacing

        # Signature line
        signature_y = 100  # Adjust this value for proper vertical placement
        pdf_canvas_obj.setFont("ArialUnicode", 12)
        pdf_canvas_obj.drawString(100, signature_y - 50,
                                  "__________________________")
        pdf_canvas_obj.drawString(100, signature_y - 70,
                                  "Подпись ООО \"Компания Мебельстайл\"")

        # Add page number to each page
        pdf_canvas_obj.setFont("ArialUnicode", 8)
        pdf_canvas_obj.drawString(
            500, 50, f'Страница {page_number + 1} из {ceil(total_entries / entries_per_page)}')

        # Add a new page if not the last page
        if page_number < (total_entries // entries_per_page):
            pdf_canvas_obj.showPage()

    try:
        pdf_canvas_obj.save()
        print(f"PDF generated: {pdf_filename}")
    except Exception as e:
        print(f"Error during PDF generation: {e}")


def on_mouse_wheel(event):
    scroll_canvas.yview_scroll(int(-1*(event.delta/120)), "units")


def on_mouse_enter(event):
    scroll_canvas.bind_all("<MouseWheel>", on_mouse_wheel)


def on_mouse_leave(event):
    scroll_canvas.unbind_all("<MouseWheel>")

# values in JSON


WORK_FILES_FOLDER = 'work_files'
if not os.path.exists(WORK_FILES_FOLDER):
    os.makedirs(WORK_FILES_FOLDER)


def save_values_to_file(file_path, values):
    file_path_in_work_files = os.path.join(WORK_FILES_FOLDER, file_path)
    with open(file_path_in_work_files, 'w') as file:
        json.dump(values, file)


def load_values_from_file(file_path='saved_values1.json'):
    file_path_in_work_files = os.path.join(WORK_FILES_FOLDER, file_path)
    try:
        with open(file_path_in_work_files, 'r') as file:
            return json.load(file)
    except FileNotFoundError:
        return {}


def clear_default_value(event, entry_var, default_value):
    if entry_var.get() == default_value:
        entry_var.set("")


def on_exit(window):
    # Clean up or perform any necessary actions before exiting
    window.destroy()

# window options


def on_resize(event):
    # # Get the new size of the window
    # new_width = scroll_canvas.winfo_width()
    # new_height = scroll_canvas.winfo_height()

    # # Calculate the scale factor
    # scale_factor = min(new_width / initial_width, new_height / initial_height)

    # # Update the scale factor and redraw the content
    # scroll_canvas.scale("all", 0, 0, scale_factor, scale_factor)
    # # Add a short delay (e.g., 100 milliseconds) before updating the content
    window.after(500, update_after_resize)


def update_after_resize():
    # Get the new size of the window
    new_width = scroll_canvas.winfo_width()
    new_height = scroll_canvas.winfo_height()

    # Calculate the scale factor
    scale_factor = min(new_width / initial_width, new_height / initial_height)

    # Update the scale factor and redraw the content
    scroll_canvas.scale("all", 0, 0, scale_factor, scale_factor)


# Create the main window
window = tk.Tk()
window.title('-=Теплоснабжение=-')
window.geometry('1400x500+300+100')
window.iconbitmap('app_files/favicon.ico')
window.bind("<Configure>", on_resize)
# Set the initial size of the content
initial_width = 1400
initial_height = 500

# Create a Canvas with Scrollbar
scroll_canvas = tk.Canvas(window)
scroll_canvas.pack(side="left", fill="both", expand=True)

# Add a Vertical Scrollbar
scrollbar = ttk.Scrollbar(window, orient="vertical",
                          command=scroll_canvas.yview)
scrollbar.pack(side="right", fill="y")
scroll_canvas.configure(yscrollcommand=scrollbar.set)

# Create a frame inside the canvas for top content
top_frame_inside_canvas = ttk.Frame(scroll_canvas, borderwidth=0)
scroll_canvas.create_window(
    (0, 0), window=top_frame_inside_canvas, anchor='nw')

# top 1 frame
frame_top1 = ttk.Frame(top_frame_inside_canvas)
label1 = tk.Label(frame_top1,
                  text=f'Показания - Теплоснабжение - {
                      datetime.now().strftime("%B %Y")}',
                  font='Arial 24 bold',
                  background='#D30000')
label1.pack(fill='both')
frame_top1.pack(side='top', fill='both', expand=True)

# top 2 frame
frame_top2 = ttk.Frame(top_frame_inside_canvas)
label2 = tk.Label(frame_top2, text='Введите текущие показания счетчиков',
                  font='Arial 14 bold', background='grey')
label2.pack(fill='both')
frame_top2.pack(side='top', fill='both', expand=True)

# Separator top
# styling
styl = ttk.Style()
styl.configure('TSeparator', background='grey')

# separator
ttk.Separator(
    master=top_frame_inside_canvas,
    style='TSeparator',
    takefocus=1,
    cursor='plus'
).pack(fill='x', pady=10, expand=True)

# middle frame
# Load data from Excel file using the load_data function
data = load_data('res/счетчики_тепло.xlsx')

frame_middle = ttk.Frame(top_frame_inside_canvas)
entry_vars = []
entry2_vars = []
entry3_vars = []
column_values_3 = []  # To store values from the third column
entry_widgets = []  # To store entry widgets for disabling later
saved_values1 = load_values_from_file('saved_values1.json')
saved_values2 = load_values_from_file('saved_values2.json')
saved_values3 = load_values_from_file('saved_values3.json')

for i, (value1, value2, value3, value4) in enumerate(data):
    create_widgets(frame_middle, i, value1, value2, value3,
                   value4, saved_values1, saved_values2, saved_values3)

frame_middle.pack(side='top', fill='both', expand=True)

# Separator bottom

# separator
ttk.Separator(
    master=top_frame_inside_canvas,
    style='TSeparator',
    takefocus=1,
    cursor='plus'
).pack(fill='x', padx=10, expand=True)

# bottom frame
frame_bottom = ttk.Frame(top_frame_inside_canvas)
label4 = tk.Label(
    frame_bottom, text='После ввода данных нажмите отправить.\n После этого проверьте папку с программой там должен быть файл с отчетом.', font='Arial 9 bold')
label4.pack()
frame_bottom.pack(side='top', fill='both', expand=True)
button_style = ttk.Style()
button_style.configure('Submit.TButton', background='blue',
                       foreground='black', font='Arial 14 bold')
submit_button = ttk.Button(
    frame_bottom, text="Отправить", command=on_submit, style='Submit.TButton')
submit_button.pack(ipady=20, ipadx=50)
# Initialize button state
button_state = tk.StringVar(value='submit')

# Reload values
frame_reload_excel = ttk.Frame(top_frame_inside_canvas)
frame_reload_excel.pack(side='bottom', fill='x',
                        expand=True)  # Updated pack method

label5 = tk.Label(frame_reload_excel,
                  text='\n\n\nПри изменении списка счетчиков \n нажмите обновить список перед заполнением')
label5.grid(row=0, column=0, sticky='we')

reload_button = ttk.Button(
    frame_reload_excel, text='Обновить список счетчиков', command=update_gui)
reload_button.grid(row=1, column=0, sticky='we')

update_gui()


# Bind events for mouse over the window
scroll_canvas.bind('<Enter>', on_mouse_enter)
scroll_canvas.bind('<Leave>', on_mouse_leave)

# Bind on_exit to the close event of the window
window.protocol("WM_DELETE_WINDOW", lambda: on_exit(window))

# Update the scroll region
scroll_canvas.update_idletasks()
scroll_canvas.config(scrollregion=scroll_canvas.bbox("all"))

# Bind on_exit to the close event of the window
window.protocol("WM_DELETE_WINDOW", lambda: on_exit(window))
# Call on_resize initially to set the initial window scale
on_resize(tk.Event())
# Run the Tkinter event loop
window.mainloop()
